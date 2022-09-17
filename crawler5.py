import requests
import lxml
from bs4 import BeautifulSoup, Comment
from collections import defaultdict
import bs4
from collections import OrderedDict
import json
from bs4.element import Tag
import copy
from openpyxl import Workbook


wb = Workbook()

# grab the active worksheet
ws_groups = wb.active
ws_groups.title = "Groups"



url = "https://attack.mitre.org/groups/"

headers = {
  'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE'
}
f = requests.get(url, headers = headers)

groups_list = []
soup = BeautifulSoup(f.content,'lxml')

# para obtener los links y hacerlo recursivo
groups = soup.find('table',{'class':'table table-bordered table-alternate mt-2'}).find_all('a')

# para obtener la informacion de los grupos
glist = soup.find('table',{'class': 'table table-bordered table-alternate mt-2'}).tbody.find_all('tr')

ws_groups['A1']='ID'
ws_groups['B1']='Name'
ws_groups['C1']='Associated Groups'
ws_groups['D1']='Description'
ws_groups['E1']='Created'


gcont=2

for g in glist:
    g_data=[]
    gcols = g.find_all('td')
#    print(gcols)
    for gcol in gcols:
        g_data.append(gcol.text.strip())

    ws_groups['A'+str(gcont)]=g_data[0]
    ws_groups['B'+str(gcont)]=g_data[1]
    ws_groups['C'+str(gcont)]=g_data[2]
    ws_groups['D'+str(gcont)]=g_data[3]
    gcont=gcont+1



num = 0

data=[]

#print(groups)

for anchor in groups:
  #print(anchor['href'])
  urls = 'https://attack.mitre.org' + anchor['href']
  if ('software' not in anchor['href']) and (urls not in groups_list):
    groups_list.append(urls)

group_info=[]


ws_tech_by_group = wb.create_sheet("Tech_by_group")
ws_tech_by_group['A1']='Group'
ws_tech_by_group['B1']='Domain'
ws_tech_by_group['C1']='ID'
ws_tech_by_group['D1']='SubID'
ws_tech_by_group['E1']='Name'
ws_tech_by_group['F1']='Use'


ws_sw_by_group = wb.create_sheet("SW_by_group")
ws_sw_by_group['A1']='Group'
ws_sw_by_group['B1']='ID'
ws_sw_by_group['C1']='Name'
ws_sw_by_group['D1']='References'
ws_sw_by_group['E1']='Techniques'

cont=2
contSW=2
contlink=2

for link in groups_list:
    try:
        group_f_date = requests.get(link, headers = headers)
        group_soup_date = BeautifulSoup(group_f_date.content, 'lxml')
        year = group_soup_date.find('div',{'class' : 'card-body'}).find_all(class_='row card-data')
        for i in range(len(year)):
            if year[i].text.find("Created")>0:
                ws_groups['E'+str(contlink)]=year[i].text.strip()
                contlink=contlink+1
                break

    except Exception:
        pass

for link in groups_list:
    try:
        group_f = requests.get(link, headers = headers)
        group_soup = BeautifulSoup(group_f.content, 'lxml')

        group_Techniques = group_soup.find('table',{'class': 'table techniques-used background table-bordered'}).find_all('tr')  #Validar si hay tabla de Techniques

#        year = group_soup.find('div',{'class' : 'card-body'}).find_all(class_='row card-data')
#        if (len(year)==5):
#            ws_groups['E'+str(contlink)]=year[3].text.strip()
#        elif (len(year)==6):
#            ws_groups['E'+str(contlink)]=year[4].text.strip()
#        elif (len(year)==4):
#            ws_groups['E'+str(contlink)]=year[2].text.strip()
#        else:

#    if isinstance(group_Techniques,Tag):

        for row in group_Techniques:

            cols = row.find_all('td')

            table_data=[]
            for col in cols:
                table_data.append(col.text.strip())

            if (len(table_data)==4):
                ws_tech_by_group['A'+str(cont)]=link
                ws_tech_by_group['B'+str(cont)]=table_data[0]
                ws_tech_by_group['C'+str(cont)]=table_data[1]
                ws_tech_by_group['D'+str(cont)]=''
                ws_tech_by_group['E'+str(cont)]=table_data[2]
                ws_tech_by_group['F'+str(cont)]=table_data[3]

                cont=cont+1

            if (len(table_data)==5):
                ws_tech_by_group['A'+ str(cont)]=link
                ws_tech_by_group['B'+str(cont)]=table_data[0]
                ws_tech_by_group['C'+str(cont)]=table_data[1]
                ws_tech_by_group['D'+str(cont)]=table_data[2]
                ws_tech_by_group['E'+str(cont)]=table_data[3]
                ws_tech_by_group['F'+str(cont)]=table_data[4]
                cont=cont+1

        group_Software = group_soup.find('table',{'class': 'table table-bordered table-alternate mt-2'}).tbody.find_all('tr')
#        print(group_Software)

        for row_sw in group_Software:
#        for i in range(len(group_Software)-1):
#            print(group_Software[2])
#            print(row_sw)
            cols_sw = row_sw.find_all('td')

            table_data_sw=[]

            for col_sw in cols_sw:

                table_data_sw.append(col_sw.text.strip())

            ws_sw_by_group['A'+str(contSW)]=link
            ws_sw_by_group['B'+str(contSW)]=table_data_sw[0]
            ws_sw_by_group['C'+str(contSW)]=table_data_sw[1]
            ws_sw_by_group['D'+str(contSW)]=table_data_sw[2]
            ws_sw_by_group['E'+str(contSW)]=table_data_sw[3]
            contSW=contSW+1


    except Exception:
        pass

###################################### TECHNIQUES #######################################################

url_tech = "https://attack.mitre.org/techniques/enterprise/"

f_tech = requests.get(url_tech, headers = headers)

tech_list = []
soup_tech = BeautifulSoup(f_tech.content,'lxml')

techs = soup_tech.find('table',{'class':'table-techniques'}).find_all('a')

techs_info = soup_tech.find('table',{'class':'table-techniques'}).tbody.find_all('tr')

data=[]


for anchor in techs:
  #print(anchor['href'])
  urls = 'https://attack.mitre.org' + anchor['href']
  if (urls not in tech_list):
    tech_list.append(urls)


ws_tech = wb.create_sheet("techniques")
ws_tech['A1']='ID'
ws_tech['B1']='SubID'
ws_tech['C1']='Name'
ws_tech['D1']='Description'

cont_tech=2
tech_temp=""

for technique in techs_info:
    try:
#        print(technique.get('class')[0])
        if (technique.get('class')[0]=='technique'):

            cols_tech = technique.find_all('td')

            table_data_tech=[]

            for col_tech in cols_tech:
                table_data_tech.append(col_tech.text.strip())

            ws_tech['A'+str(cont_tech)]=table_data_tech[0]
            ws_tech['B'+str(cont_tech)]=''
            ws_tech['C'+str(cont_tech)]=table_data_tech[1]
            ws_tech['D'+str(cont_tech)]=table_data_tech[2]
            cont_tech=cont_tech+1
            tech_temp=table_data_tech[0]

        if (technique.get('class')[0]=='sub'):

            cols_tech = technique.find_all('td')

            table_data_tech=[]

            for col_tech in cols_tech:
                table_data_tech.append(col_tech.text.strip())

            ws_tech['A'+str(cont_tech)]=tech_temp
            ws_tech['B'+str(cont_tech)]=table_data_tech[1]
            ws_tech['C'+str(cont_tech)]=table_data_tech[2]
            ws_tech['D'+str(cont_tech)]=table_data_tech[3]
            cont_tech=cont_tech+1

    except Exception:
        pass

################################################# CONTROLS BY TECH ###################################################################################


ws_mit_by_tech = wb.create_sheet("mit_by_tech")
ws_mit_by_tech['A1']='Technique'
ws_mit_by_tech['B1']='ID'
ws_mit_by_tech['C1']='Mitigation'
ws_mit_by_tech['D1']='Description'

cont_mit=2


ws_det_by_tech = wb.create_sheet("det_by_tech")
ws_det_by_tech['A1']='Technique'
ws_det_by_tech['B1']='ID'
ws_det_by_tech['C1']='Data Source'
ws_det_by_tech['D1']='Description'

cont_det=2

for link in tech_list:
    try:
        control_f = requests.get(link, headers = headers)
        control_soup = BeautifulSoup(control_f.content, 'lxml')

        flag=control_soup.find('h2',{'class':'pt-3', 'id':'examples'})
        #        print(flag)
        if (flag!=None):
            mit_info = control_soup.find_all('table',{'class':'table table-bordered table-alternate mt-2'})[1].tbody.find_all('tr')

        else:
            mit_info = control_soup.find_all('table',{'class':'table table-bordered table-alternate mt-2'})[0].tbody.find_all('tr')
#        print(mit_info)

        for mitigation in mit_info:

            cols_mit = mitigation.find_all('td')
            #print(cols_mit)
            table_data_mit=[]

            for col_mit in cols_mit:
                table_data_mit.append(col_mit.text.strip())

            ws_mit_by_tech['A'+str(cont_mit)]=link
            ws_mit_by_tech['B'+str(cont_mit)]=table_data_mit[0]
            ws_mit_by_tech['C'+str(cont_mit)]=table_data_mit[1]
            ws_mit_by_tech['D'+str(cont_mit)]=table_data_mit[2]
            cont_mit=cont_mit+1

        det_info = control_soup.find('table',{'class':'table datasources-table table-bordered'}).tbody.find_all('tr')
        ID_temp=""
        DS_temp=""

        for detection in det_info:

            cols_det = detection.find_all('td')
            #print(cols_mit)
            table_data_det=[]

            for col_det in cols_det:
                table_data_det.append(col_det.text.strip())

            if (table_data_det[0]==""):
                ws_det_by_tech['A'+str(cont_det)]=link
                ws_det_by_tech['B'+str(cont_det)]=ID_temp
                ws_det_by_tech['C'+str(cont_det)]=DS_temp
                ws_det_by_tech['D'+str(cont_det)]=table_data_det[2]
                cont_det=cont_det+1
            else:
                ws_det_by_tech['A'+str(cont_det)]=link
                ws_det_by_tech['B'+str(cont_det)]=table_data_det[0]
                ws_det_by_tech['C'+str(cont_det)]=table_data_det[1]
                ws_det_by_tech['D'+str(cont_det)]=table_data_det[2]
                cont_det=cont_det+1
                ID_temp=table_data_det[0]
                DS_temp=table_data_det[1]

    except Exception:
        pass

######################################## DATA SOURCE LIST #################################################################################################


url_ds = "https://attack.mitre.org/datasources/"

f_ds = requests.get(url_ds, headers = headers)

ds_list = []
soup_ds = BeautifulSoup(f_ds.content,'lxml')

ds_info = soup_ds.find('table',{'table table-bordered table-alternate mt-2'}).tbody.find_all('tr')

data=[]


ws_ds = wb.create_sheet("data sources")
ws_ds['A1']='ID'
ws_ds['B1']='Name'
ws_ds['C1']='Description'

cont_ds=2

for ds in ds_info:
    try:
        cols_ds = ds.find_all('td')
        table_data_ds=[]
        for col_ds in cols_ds:

            table_data_ds.append(col_ds.text.strip())

        ws_ds['A'+str(cont_ds)]=table_data_ds[0]
        ws_ds['B'+str(cont_ds)]=table_data_ds[1]
        ws_ds['C'+str(cont_ds)]=table_data_ds[2]

        cont_ds=cont_ds+1

    except Exception:
        pass


######################################## MITIGATION LIST #################################################################################################

url_mitigation = "https://attack.mitre.org/mitigations/enterprise/"

f_mitigation = requests.get(url_mitigation, headers = headers)

soup_mitigation = BeautifulSoup(f_mitigation.content,'lxml')

mitigation_info = soup_mitigation.find('table',{'table table-bordered table-alternate mt-2'}).tbody.find_all('tr')


ws_mitigation = wb.create_sheet("mitigations")
ws_mitigation['A1']='ID'
ws_mitigation['B1']='Name'
ws_mitigation['C1']='Description'

cont_mitigation=2

for mitigation in mitigation_info:
    try:
        cols_mitigation = mitigation.find_all('td')
        table_data_mitigation=[]

        for col_mitigation in cols_mitigation:
            table_data_mitigation.append(col_mitigation.text.strip())

        ws_mitigation['A'+str(cont_mitigation)]=table_data_mitigation[0]
        ws_mitigation['B'+str(cont_mitigation)]=table_data_mitigation[1]
        ws_mitigation['C'+str(cont_mitigation)]=table_data_mitigation[2]

        cont_mitigation=cont_mitigation+1

    except Exception:
        pass


wb.save("sample7.xlsx")
